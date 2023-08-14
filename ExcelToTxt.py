# import openpyxl
# import re
#
# # ÊâìÂºÄexcel
# workbook = openpyxl.load_workbook('file.xlsx', data_only=True)
# sheet_names = workbook.sheetnames
#
# # ËØªÂèñÊâÄÊúâsheet
# for sheet_name in sheet_names:
#     # Select the sheet
#     sheet = workbook[sheet_name]
#
#     # Get the maximum row and column count
#     max_row = sheet.max_row
#     max_column = sheet.max_column
#
#     # Loop through each cell
#     for row in range(1, max_row + 1):
#         for column in range(1, max_column + 1):
#             # Get the cell value
#             cell_value = sheet.cell(row=row, column=column).value
#             # Check if the cell is not empty
#             if cell_value is not None:
#                 cell_value = cell_value.rstrip('\n')
#                 # ÂéªÊéâÁ©∫Ê†º
#                 cell_value = re.sub("(?:\u0020|\u3000|\u00A0|\u2002|\u2003|\s+|_x000D_)", "", cell_value)
#
#                 # Ëã±ÊñáÂ∞èÂÜô
#                 # cell_value = cell_value.lower()
#                 # ÂéªÈô§Á¨¶Âè∑
#                 symbols = [",", "!", "?", ";", ":", "Ôºå", "Ôºõ", ".", "Ôºü", "Ôºõ", "Ôºå", "„ÄÅ"]
#                 for symbol in symbols:
#                     cell_value = cell_value.replace(symbol, "„ÄÇ")
#
#                 # Write the cell value to a txt file
#                 with open('output.txt', 'a+', encoding='utf-8') as file:
#                     if cell_value not in file:
#                         file.write(cell_value + '\n')


import openpyxl
import re


def ExcelToTxt(excelname, txtname):
    # ÊâìÂºÄexcel
    workbook = openpyxl.load_workbook(excelname, data_only=True)
    sheet_names = workbook.sheetnames
    # ÂàõÂª∫‰∏Ä‰∏™Á©∫ÁöÑÈõÜÂêà
    written_values = set()
    # ËØªÂèñÊâÄÊúâsheet
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]

        # Get the maximum row and column count
        max_row = sheet.max_row
        max_column = sheet.max_column

        # Loop through each cell
        for row in range(1, max_row + 1):
            for column in range(1, max_column + 1):
                # Get the cell value
                cell_value = sheet.cell(row=row, column=column).value
                # Â¶ÇÊûúÂçïÂÖÉÊ†º‰∏ç‰∏∫Á©∫
                if cell_value is not None:
                    cell_value = cell_value.rstrip('\n')
                    # ÂéªÊéâÁ©∫Ê†º
                    cell_value = re.sub("(?:\u0020|\u3000|\u00A0|\u2002|\u2003|\s+|_x000D_)", "", cell_value)
                    # Ëã±ÊñáÂ∞èÂÜô
                    cell_value = cell_value.lower()
                    # ÂéªÈô§Á¨¶Âè∑
                    symbols = ["‚Äî", "(", "%", "‚óè", "‚àö", ")", "‚Ññ", "‚úö", "Ôπ£", "/", "Ôºù", "\"", "Ôºè", "\'", "Ôºâ", "Ôºé", "„éù",
                               "ÓÑª", "‚Äù", "+", "‚â§", "-", "Ôºá", "‚àè", "[", "ÔºÉ", "‚âà", "‚Äò", "*", "#", "‚Äú", "", "‚Ä≤", "Ôºä", "Ôºà",
                               "ÔΩû", "ÓÇÉ", "‚Äô", "‚àΩ", "„é°", "√∑", "¬∞", "‚Ä¶", "„Äë", "|", "Ôºç", "¬±", "‚ÑÉ", "ÔºÖ", "&", "@", "√ó", "‚ñ°",
                               "¬∑", "Ôºö", "‚Äñ", "‚â•", "^", "Ôºú", "„Çú", "Ôºã", "„éú", "=", ">", "`", "]", "ÔºÅ", "Ôºû", "„Äê", "Ôπ§", "~",
                               "„ÄÇ", "<", "!", "?", ";", ":", "Ôºå", "Ôºõ", ".", "Ôºü", "Ôºõ", "Ôºå", "„ÄÅ", "\\", "„Äã", "‚Äπ", "‚úñ","Ô∏è",
                               "‚ñ†", "‚ñ≥", "‚ôÄ", "‚óã", "¬ù", "}", "‚Üë", "Ô∏ø", "„ÄÇ", ",", "ÔπÄ", "‚î≥", "„Ää", "‚ñΩ", "Óó•", "Óâ£", "ÔºÇ", "‚Ä∫",
                               "Ôπë", "‚î∫", "ÔΩù", "{", "„Äà", "‚Äª", "Óèá", "ÔΩõ", "‚ñè", ",", "¬°", "ÔºÇ", "„Ää", "¬£", "¬∏", "Ôπë", "¬ù", ",",
                               "Ó°¢", "Ôºº", "Ô∏è", "„Äã", "¬´", "‚ñ†", "¬®", "‚Üë", "Ôπ¢", "¬Ø", "¬ª", "{", "ÔøΩ", "¬¥", "}", "‚ùå", "‚úñ", "üëÜ",
                               "¬¨", "‚Üí"]
                    for symbol in symbols:
                        cell_value = cell_value.replace(symbol, "„ÄÇ")
                    # Âà†Èô§ËøûÁª≠ÁöÑÂè•Âè∑
                    while "„ÄÇ„ÄÇ" in cell_value:
                        cell_value = cell_value.replace("„ÄÇ„ÄÇ", "„ÄÇ")
                    # ÂÜôÂÖ•txt
                    with open(txtname, 'a+', encoding='utf-8') as file:
                        # Ê£ÄÊü•ÂΩìÂâçÂÄºÊòØÂê¶Âú®ÈõÜÂêà‰∏≠
                        if cell_value not in written_values:
                            # Â¶ÇÊûú‰∏çÂú®ÔºåÂ∞±ÂÜôÂÖ•Âπ∂Ê∑ªÂä†Âà∞ÈõÜÂêà‰∏≠
                            file.write(cell_value + '\n')
                            written_values.add(cell_value)
                        else:
                            # Â¶ÇÊûúÂú®ÔºåÂ∞±Ë∑≥Ëøá
                            pass


ExcelToTxt('ÂâØÊú¨MRÊ£ÄÊü•.xlsx', 'MRData.txt')
ExcelToTxt('ÂâØÊú¨XÁ∫øÊ£ÄÊü•Êï∞ÊçÆ.xlsx', 'XrayData.txt')
ExcelToTxt('ÂâØÊú¨CTÊ£ÄÊü•‰ø°ÊÅØ.xlsx', 'CTData.txt')
